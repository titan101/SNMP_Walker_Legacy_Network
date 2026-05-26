from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .discovery import DiscoveryResult, EXPORT_COLUMNS, MIB_WALK_PLAN, topology_context_lines, topology_edges


MIB_PLAN_COLUMNS = ["mode", "operation", "mib", "name", "oid", "collects"]
WALK_STATUS_COLUMNS = ["ip", "hostname", "mode", "mib", "name", "oid", "operation", "rows", "status", "error"]
INTERFACE_COLUMNS = ["ip", "hostname", "if_index", "description", "alias", "admin_status", "oper_status", "speed_mbps"]
ENTITY_COLUMNS = ["ip", "hostname", "entity_index", "description", "model", "serial"]
NEIGHBOR_COLUMNS = [
    "ip",
    "hostname",
    "protocol",
    "local_port",
    "remote_device",
    "remote_port",
    "remote_platform",
    "link_type",
    "bandwidth_mbps",
    "table_index",
]
TOPOLOGY_COLUMNS = [
    "source",
    "source_ip",
    "source_port",
    "target",
    "target_port",
    "protocol",
    "remote_platform",
    "link_type",
    "bandwidth_mbps",
]
TOPOLOGY_CONTEXT_COLUMNS = ["line"]


def results_to_csv(results: list[DiscoveryResult], include_community: bool = False) -> bytes:
    buffer = io.StringIO(newline="")
    columns = export_columns(include_community)
    writer = csv.DictWriter(buffer, fieldnames=columns)
    writer.writeheader()
    for result in results:
        data = result.public_dict(include_community=include_community)
        writer.writerow({column: data.get(column, "") for column in columns})
    return buffer.getvalue().encode("utf-8-sig")


def results_to_xlsx(results: list[DiscoveryResult], include_community: bool = False) -> bytes:
    payload = io.BytesIO()
    workbook = Workbook()
    write_sheet(
        workbook.active,
        "Devices",
        [result.public_dict(include_community=include_community) for result in results],
        export_columns(include_community),
    )
    write_sheet(workbook.create_sheet(), "MIB Walk Plan", MIB_WALK_PLAN, MIB_PLAN_COLUMNS)
    write_sheet(workbook.create_sheet(), "Walk Status", flatten_detail_rows(results, "walk_errors"), WALK_STATUS_COLUMNS)
    write_sheet(workbook.create_sheet(), "Interfaces", flatten_detail_rows(results, "interface_rows"), INTERFACE_COLUMNS)
    write_sheet(workbook.create_sheet(), "Entities", flatten_detail_rows(results, "entity_rows"), ENTITY_COLUMNS)
    write_sheet(workbook.create_sheet(), "Neighbors", flatten_detail_rows(results, "neighbor_rows"), NEIGHBOR_COLUMNS)
    write_sheet(workbook.create_sheet(), "Topology", topology_edges(results), TOPOLOGY_COLUMNS)
    write_sheet(workbook.create_sheet(), "Topology Context", topology_context_lines(results), TOPOLOGY_CONTEXT_COLUMNS)

    workbook.save(payload)
    payload.seek(0)
    return payload.getvalue()


def export_columns(include_community: bool = False) -> list[str]:
    columns = list(EXPORT_COLUMNS) + ["snmp_status", "snmp_error"]
    if include_community:
        columns.append("matched_community")
    return columns


def flatten_detail_rows(results: list[DiscoveryResult], field_name: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for result in results:
        detail_rows = getattr(result, field_name)
        rows.extend(detail_rows)
    return rows


def write_sheet(worksheet: Worksheet, sheet_name: str, rows: list[dict[str, str]], columns: list[str]) -> None:
    worksheet.title = sheet_name
    worksheet.append(columns)
    for row in rows:
        worksheet.append([safe_cell_value(row.get(column, "")) for column in columns])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max((len(value) for value in values), default=10) + 2, 48)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def safe_cell_value(value) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
