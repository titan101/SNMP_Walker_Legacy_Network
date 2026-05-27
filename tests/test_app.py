import io
import time

from openpyxl import load_workbook

from snmp_walker.discovery import DiscoveryResult
from snmp_walker.web import app
from snmp_walker.web import deserialize_results


def test_index_get_loads():
    client = app.test_client()
    response = client.get("/")
    assert response.status_code == 200
    assert b"SNMP Discovery" in response.data
    assert b"MIB/OID Coverage" in response.data
    assert b"sysDescr.0" in response.data
    assert b"live-results-body" in response.data
    assert b"remember-inputs" in response.data
    assert b"snmpWalker.communities" in response.data
    assert b'name="selected_oids"' in response.data


def test_index_post_runs_fast_identity_scan_by_default(monkeypatch):
    captured = {}

    def fake_discover_many(targets, communities, **kwargs):
        captured["targets"] = targets
        captured["communities"] = communities
        captured["kwargs"] = kwargs
        return [
            DiscoveryResult(
                ip="192.0.2.1",
                pingable_or_not="yes",
                hostname="legacy-rtr",
                device_model="ASR9000",
                device_type="router",
                software_version="7.5.2",
                snmp_status="ok",
            )
        ]

    monkeypatch.setattr("snmp_walker.web.discover_many", fake_discover_many)
    client = app.test_client()
    response = client.post(
        "/",
        data={
            "targets": "192.0.2.1",
            "communities": "public",
            "max_hosts": "10",
            "workers": "1",
            "ping_timeout_ms": "200",
            "snmp_timeout_seconds": "0.3",
            "snmp_retries": "0",
            "do_ping": "on",
        },
    )
    assert response.status_code == 200
    assert b"legacy-rtr" in response.data
    assert captured["targets"] == ["192.0.2.1"]
    assert captured["communities"] == ["public"]
    assert captured["kwargs"]["walk_details"] is False
    assert captured["kwargs"]["walk_traffic_tables"] is False
    assert "1.3.6.1.2.1.1.1.0" in captured["kwargs"]["selected_oids"]


def test_index_post_passes_selected_oids(monkeypatch):
    captured = {}

    def fake_discover_many(targets, communities, **kwargs):
        captured["kwargs"] = kwargs
        return [DiscoveryResult(ip="192.0.2.1", snmp_status="failed")]

    monkeypatch.setattr("snmp_walker.web.discover_many", fake_discover_many)
    client = app.test_client()
    response = client.post(
        "/",
        data={
            "targets": "192.0.2.1",
            "communities": "public",
            "oid_selection_present": "1",
            "selected_oids": ["1.3.6.1.2.1.1.1.0", "1.0.8802.1.1.2.1.4.1.1.9"],
        },
    )
    assert response.status_code == 200
    assert captured["kwargs"]["selected_oids"] == [
        "1.3.6.1.2.1.1.1.0",
        "1.0.8802.1.1.2.1.4.1.1.9",
    ]


def test_index_post_requires_one_base_oid():
    client = app.test_client()
    response = client.post(
        "/",
        data={
            "targets": "192.0.2.1",
            "communities": "public",
            "oid_selection_present": "1",
            "selected_oids": "1.0.8802.1.1.2.1.4.1.1.9",
        },
    )
    assert response.status_code == 200
    assert b"Select at least one base GET OID" in response.data


def test_index_post_renders_topology_when_neighbors_exist(monkeypatch):
    def fake_discover_many(targets, communities, **kwargs):
        return [
            DiscoveryResult(
                ip="192.0.2.1",
                hostname="r1",
                snmp_status="ok",
                neighbor_rows=[
                    {
                        "ip": "192.0.2.1",
                        "hostname": "r1",
                        "protocol": "LLDP",
                        "local_port": "localPort 1",
                        "remote_device": "sw1",
                        "remote_port": "Gi1/0/1",
                    }
                ],
            )
        ]

    monkeypatch.setattr("snmp_walker.web.discover_many", fake_discover_many)
    client = app.test_client()
    response = client.post("/", data={"targets": "192.0.2.1", "communities": "public", "walk_details": "on"})
    assert response.status_code == 200
    assert b"Topology" in response.data
    assert b"sw1" in response.data


def test_index_post_shows_no_snmp_diagnostic(monkeypatch):
    def fake_discover_many(targets, communities, **kwargs):
        return [
            DiscoveryResult(
                ip="192.0.2.1",
                pingable_or_not="no",
                snmp_status="failed",
                snmp_error="No SNMP response received before timeout",
            )
        ]

    monkeypatch.setattr("snmp_walker.web.discover_many", fake_discover_many)
    client = app.test_client()
    response = client.post("/", data={"targets": "192.0.2.1", "communities": "public", "do_ping": "on"})
    assert response.status_code == 200
    assert b"No SNMP responders were found" in response.data
    assert b"No targets answered ICMP ping" in response.data


def test_scan_api_reports_progress_and_result_page(monkeypatch):
    def fake_discover_iter(targets, communities, **kwargs):
        for target in targets:
            yield DiscoveryResult(ip=target, hostname=f"host-{target}", snmp_status="ok")

    monkeypatch.setattr("snmp_walker.web.discover_iter", fake_discover_iter)
    client = app.test_client()
    response = client.post(
        "/api/scans",
        data={
            "targets": "192.0.2.1\n192.0.2.2",
            "communities": "public",
            "workers": "2",
        },
    )
    assert response.status_code == 200
    job_id = response.get_json()["job_id"]

    payload = {}
    for _ in range(20):
        status = client.get(f"/api/scans/{job_id}")
        assert status.status_code == 200
        payload = status.get_json()
        if payload["status"] == "done":
            break
        time.sleep(0.05)

    assert payload["status"] == "done"
    assert payload["completed"] == 2
    assert payload["total"] == 2
    assert payload["results"][0]["ip"] == "192.0.2.1"

    result_page = client.get(f"/scan/{job_id}/results")
    assert result_page.status_code == 200
    assert b"host-192.0.2.1" in result_page.data


def test_scan_api_returns_validation_errors():
    client = app.test_client()
    response = client.post("/api/scans", data={"targets": "", "communities": "public"})
    assert response.status_code == 400
    assert b"Add at least one IP" in response.data


def test_download_csv_from_result_payload():
    client = app.test_client()
    response = client.post(
        "/download/csv",
        data={
            "results_json": '[{"ip":"192.0.2.1","hostname":"r1","snmp_status":"ok"}]',
        },
    )
    assert response.status_code == 200
    assert response.headers["Content-Type"].startswith("text/csv")
    assert b"192.0.2.1" in response.data


def test_download_xlsx_includes_walk_detail_sheets():
    client = app.test_client()
    response = client.post(
        "/download/xlsx",
        data={
            "results_json": (
                '[{"ip":"192.0.2.1","hostname":"r1","snmp_status":"ok",'
                '"interface_rows":[{"ip":"192.0.2.1","hostname":"r1","if_index":"1","description":"Gi0/0","oper_status":"up"}],'
                '"entity_rows":[{"ip":"192.0.2.1","hostname":"r1","entity_index":"1","model":"ASR9000","serial":"ABC12345"}],'
                '"neighbor_rows":[{"ip":"192.0.2.1","hostname":"r1","protocol":"LLDP","local_port":"localPort 1","remote_device":"sw1","remote_port":"Gi1/0/1"}],'
                '"walk_errors":[{"ip":"192.0.2.1","hostname":"r1","mode":"inventory","mib":"IF-MIB","name":"ifDescr","oid":"1.3.6.1.2.1.2.2.1.2","operation":"WALK","rows":"1","status":"ok","error":""}]}]'
            ),
        },
    )
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data), read_only=True)
    assert workbook.sheetnames == [
        "Devices",
        "MIB Walk Plan",
        "Walk Status",
        "Interfaces",
        "Entities",
        "Neighbors",
        "Topology",
        "Topology Context",
    ]
    assert workbook["MIB Walk Plan"]["D2"].value == "sysDescr.0"
    assert workbook["Walk Status"]["I2"].value == "ok"
    assert workbook["Interfaces"]["D2"].value == "Gi0/0"
    assert workbook["Entities"]["F2"].value == "ABC12345"
    assert workbook["Topology"]["D2"].value == "sw1"
    assert workbook["Topology Context"]["A2"].value == "r1 [localPort 1] <-> [Gi1/0/1] sw1 (LLDP)"


def test_deserialize_results_ignores_bad_download_payload_rows():
    rows = deserialize_results(
        '[{"ip":"192.0.2.1","hostname":"r1","extra":"ignored"},'
        '{"hostname":"missing-ip"},'
        '"bad-row"]'
    )
    assert len(rows) == 1
    assert rows[0].ip == "192.0.2.1"
    assert rows[0].hostname == "r1"


def test_deserialize_results_preserves_walk_detail_rows():
    rows = deserialize_results(
        '[{"ip":"192.0.2.1","interface_rows":[{"if_index":1,"description":"Gi0/0"}]}]'
    )
    assert rows[0].interface_rows == [{"if_index": "1", "description": "Gi0/0"}]
